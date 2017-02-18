#!/usr/bin/env python3
import openpyxl
import sys
import csv
from collections import defaultdict

def get_company_row(sheet, row_start, col_name_n):
    company_row = []

    row = row_start
    col = col_name_n
    while sheet.cell(row=row, column=col).value:
        cell = sheet.cell(row=row, column=col)
        company_row.append(cell.value)
        row += 1
    return company_row

def get_company_col(sheet, col_start, row_name_n):
    company_col = []
    row = row_name_n
    col = col_start
    while sheet.cell(row=row, column=col).value:
        cell = sheet.cell(row=row, column=col)
        company_col.append(cell.value)
        col += 1
    return company_col

def convert_2d_to_1d(sheet, company_row,company_col, row_start, col_start):
    records = defaultdict(list)
    for row in range(row_start,len(company_row)+row_start):
        company_a = company_row[row-row_start]
        for col in range(col_start,len(company_col)+col_start):
            company_b = company_col[col-col_start]
            try:
                float(sheet.cell(row=row,column=col).value)
            except:
                continue
            cell = sheet.cell(row=row,column=col)
            value = cell.value
            if value > 0:
                records[company_a].append([
                    company_b, value
                ])
    return records

def convert_to_records(filename):
    row_start = 3
    col_name_n = 1

    col_start = 2
    row_name_n = 1
    wb = openpyxl.load_workbook(filename)
    all_records = {}
    for sheet in wb.worksheets:
        company_row = get_company_row(sheet, row_start, col_name_n)
        company_col = get_company_col(sheet, col_start, row_name_n)
        records = convert_2d_to_1d(sheet, company_row,company_col, row_start, col_start)
        all_records[sheet.title] = records
    return all_records

def group(lst, n):
    for i in range(0, len(lst), n):
        val = lst[i:i+n]
        if len(val) == n:
           yield tuple(val)

def mix_records(left_account, left_records, right_account, right_records):
    for left_com, left_com_records in left_records.items():
        for left_com_record in left_com_records:
            right_com = left_com_record[0]
            left_value = left_com_record[1]
            if right_com in right_records:
                for right_record in right_records[right_com]:
                    yield([
                        left_com, right_com, left_account, left_value,
                        right_com, right_record[0], right_account, right_record[1]
                    ])
            else:
                yield([
                        left_com, right_com, left_account, left_value,
                        right_com, None, None, None
                    ])

def mix_all(account_groups, all_records):
    for account_group in account_groups:
        for left_account, right_account in [account_group, account_group[::-1]]:
            left_records = all_records[left_account]
            right_records = all_records[right_account]
            for mix_record in mix_records(left_account, left_records, right_account, right_records):
                yield mix_record

def main(argv):
    try:
        filename = argv[0]
    except:
        print('Did you forget the filename?')

    all_records = convert_to_records(filename)
    accounts = list(all_records.keys())
    account_groups = list(group(accounts, 2))
    with open('output.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, dialect='excel')
        writer.writerows(mix_all(account_groups, all_records))
    
if __name__ == '__main__':
    main(sys.argv[1:])
